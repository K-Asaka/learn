import openpyxl as px

wbm = px.Workbook()
wsm = wbm.create_sheet("Report", 0)

with open('manuscript.txt', encoding='utf-8') as f:
    eof = False
    while eof == False:
        rline = ""  # 1行分の文字列を保持する
        while len(rline) < 16:
            letter = f.read(1)
            if letter == '' or letter is None:  # ファイルの最後に来た
                eof = True
                break
            elif letter == "\n":
                break
            else:
                rline += letter
        wsm.append([rline])
f.close()

print (wsm.max_row)
wsm.move_range("A30:A57", rows=-29, cols=2)
wsm.column_dimensions['A'].width = 32
wsm.column_dimensions['C'].width = 32
wsm.column_dimensions['B'].width = 1

wbm.save('styled.xlsx')
