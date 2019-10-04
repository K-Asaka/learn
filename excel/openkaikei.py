import openpyxl as px
from openpyxl.formula.translate import Translator

wbk = px.load_workbook('kaikei.xlsx', data_only=True)
wsk = wbk['Sheet1']

for row in wsk:
    rowstr = ""
    for cell in row:
        cellstr = "(%s, %s)" % (cell.coordinate, 'None' if cell.value is None else cell.value)
        rowstr += " %s " % cellstr
    print(rowstr)

in17 = wsk['A4:B8']
for row in in17:
    print(row[0].value, row[1].value)

new_wbk = px.Workbook()
wsk17_in = new_wbk.create_sheet("2017収入", 0)
wsk17_in.append(["項目", "金額"])
for row in wsk['A4:B8']:
    wsk17_in.append([row[i].value for i in range(len(row))])
wsk17_out = new_wbk.create_sheet("2017支出", 1)
wsk17_out.append(["項目", "金額"])
for row in wsk['D4:E11']:
    wsk17_out.append([row[i].value for i in range(len(row))])
new_wbk.save('new_kaikei.xlsx')

#wsk17_in['B6'] = '=SUM(B2:B5)'
# ファイルを再生成
#new_wbk.save('new_kaikei.xlsx')

wsk17_in['B6'] = Translator('=SUM(B4:B7)', 'B8').translate_formula('B6')
# ファイルを再生成
new_wbk.save('new_kaikei.xlsx')

good_wbk = px.Workbook()
wsk_in = good_wbk.create_sheet('収入', 0)
wsk_in.append(['項目', '2017年度', '2018年度'])
for row in wsk['A4:B8']:
    wsk_in.append([row[i].value for i in range(len(row))])
in18row = 17
for col in wsk_in.iter_cols(min_row=2, max_row=6, min_col=3, max_col=3):
    for cell in col:
        cell.value = wsk['B' + str(in18row)].value
        in18row += 1
wsk_in.move_range('C3:C5', rows=1)
wsk_in['C3'] = 0
good_wbk.save('good_kaikei.xlsx')
