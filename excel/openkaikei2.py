import openpyxl as px
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill

good_wbk2 = px.load_workbook('good_kaikei.xlsx', data_only=True)
wsk_in = good_wbk2[good_wbk2.sheetnames[0]]

for row in wsk_in['B2:B6']:
    for cell in row:
        cell.number_format = '#,##0'

for row in wsk_in.iter_rows(max_row=1, max_col=3):
    for cell in row:
        cell.border = Border(bottom=Side(style='thin'))

for row in wsk_in.iter_rows(min_row=6, max_row=6, max_col=3):
    for cell in row:
        cell.border = Border(top=Side(style='double'))

for col in wsk_in.iter_cols(max_col=1, max_row=6):
    for cell in col:
        cell.fill = PatternFill("solid", fgColor="EEEEEE")

good_wbk2.save('good_kaikei2.xlsx')
